# -*- coding: utf-8 -*-
"""Read working_ui/Watiting List (Portal).xlsx → working_ui/ELEMENTOR/MEDIOS/waiting_list_portal_data.js.

Expected export shape matches ClassForKids-style columns (row 1 = headers). The current sheet
lists children and parents; add two columns for the admin grid when you have them:

  • **desiredService** (aliases: Service, serviceWanted, programme) — service they want.
  • **slotTime** (aliases: Slot time, preferredSlot, timeSlot) — preferred slot / time window.

If those headers are missing, the portal still shows the queue with "—" in those cells.
"""
import json
import os
import re
from datetime import date, datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
_DEFAULT_XLSX = ROOT / "working_ui" / "Watiting List (Portal).xlsx"
XLSX = Path(os.environ.get("WAITING_LIST_PORTAL_XLSX", str(_DEFAULT_XLSX)))
OUT = ROOT / "working_ui" / "ELEMENTOR" / "MEDIOS" / "waiting_list_portal_data.js"

UK_MONTHS = [
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
]


def norm_key(h):
    if h is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(h).strip().lower())


def norm_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def cell_str(v):
    if v is None:
        return None
    if isinstance(v, float) and not isinstance(v, bool) and v == int(v):
        s = str(int(v))
        return s if s else None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        s = str(v).strip()
        return s if s else None
    return norm_str(v)


def display_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        d = v.date()
    elif isinstance(v, date):
        d = v
    else:
        s = str(v).strip()
        return s if s else None
    try:
        return f"{d.day} {UK_MONTHS[d.month - 1]} {d.year}"
    except (TypeError, ValueError, IndexError):
        return str(v).strip() or None


def pick_idx(header_map, *aliases):
    for a in aliases:
        k = norm_key(a)
        if k in header_map:
            return header_map[k]
    return None


def get_cell(row, idx):
    if idx is None or idx < 0:
        return None
    if idx >= len(row):
        return None
    return row[idx]


def child_name(row, header_map):
    i_fn = pick_idx(header_map, "childFirstName", "firstname", "child_first_name")
    i_ln = pick_idx(header_map, "childLastName", "lastname", "child_last_name")
    a = cell_str(get_cell(row, i_fn)) or ""
    b = cell_str(get_cell(row, i_ln)) or ""
    name = " ".join((a + " " + b).split()).strip()
    return name or None


def parent_line(row, header_map):
    i_pf = pick_idx(header_map, "parentFirstName", "parent_first_name")
    i_pl = pick_idx(header_map, "parentLastName", "parent_last_name")
    a = cell_str(get_cell(row, i_pf)) or ""
    b = cell_str(get_cell(row, i_pl)) or ""
    s = " ".join((a + " " + b).split()).strip()
    return s or "—"


def contact_bits(row, header_map):
    mob = cell_str(get_cell(row, pick_idx(header_map, "mobile", "phone", "telephone")))
    em = cell_str(get_cell(row, pick_idx(header_map, "username", "email", "parentEmail")))
    parts = [p for p in (mob, em) if p]
    return " · ".join(parts) if parts else "—"


def main():
    import openpyxl

    wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = list(header_row)
    header_map = {}
    for i, h in enumerate(headers):
        nk = norm_key(h)
        if nk:
            header_map[nk] = i

    i_city = pick_idx(header_map, "city", "town")
    i_pc = pick_idx(header_map, "postcode", "postalCode", "zip")
    i_fb = pick_idx(header_map, "firstBookingDate", "first_booking_date", "firstBooking")
    i_lb = pick_idx(header_map, "lastBookingDate", "last_booking_date", "lastBooking")
    i_svc = pick_idx(
        header_map,
        "desiredService",
        "servicewanted",
        "service",
        "programme",
        "program",
        "activitywanted",
    )
    i_slot = pick_idx(
        header_map,
        "slotTime",
        "slottime",
        "preferredSlot",
        "preferredslot",
        "timeSlot",
        "timeslot",
    )
    i_pref = pick_idx(
        header_map,
        "preferredTime",
        "preferredtime",
        "prefWhen",
        "availability",
    )

    rows_out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        pax = child_name(vals, header_map)
        if not pax:
            continue
        fb = display_date(get_cell(vals, i_fb))
        lb = display_date(get_cell(vals, i_lb))
        svc = norm_str(cell_str(get_cell(vals, i_svc))) or "—"
        slot_t = norm_str(cell_str(get_cell(vals, i_slot))) or "—"
        pref = norm_str(cell_str(get_cell(vals, i_pref))) or "—"
        city = norm_str(cell_str(get_cell(vals, i_city))) or "—"
        pc = norm_str(cell_str(get_cell(vals, i_pc))) or "—"
        loc = city if city != "—" else pc
        if city != "—" and pc != "—":
            loc = f"{city} · {pc}"
        since = fb or lb or "—"
        wid = len(rows_out)
        rows_out.append(
            {
                "id": f"wl-{wid}",
                "pax": pax,
                "service": svc,
                "slotTime": slot_t,
                "pref": pref,
                "ratio": "—",
                "loc": loc,
                "pri": "—",
                "match": "—",
                "cont": contact_bits(vals, header_map),
                "last": lb or "—",
                "since": since,
                "parentLine": parent_line(vals, header_map),
            }
        )

    meta = {
        "sourceFile": "Watiting List (Portal).xlsx",
        "sheet": ws.title,
        "exportedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%MZ"),
        "rowCount": len(rows_out),
        "optionalColumns": {
            "desiredService": i_svc is not None,
            "slotTime": i_slot is not None,
        },
    }
    payload = {"meta": meta, "rows": rows_out}
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as f:
        f.write("window.WAITING_LIST_PORTAL_SOURCE = ")
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")
    print("Wrote", OUT, "rows=", len(rows_out), "bytes=", OUT.stat().st_size)


if __name__ == "__main__":
    main()
