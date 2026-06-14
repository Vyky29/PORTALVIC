import csv
import json
import re
import unicodedata
from datetime import datetime, date
from pathlib import Path
import openpyxl

ROOT = Path(r"c:\Users\info\PORTAL")
SPREAD = ROOT / "SPREADSHEETS"
OUT = ROOT / "database"

# Term calendar closures (bank holidays / vacation) — not inferred from the timetable sheet.
TERM_BREAK_FROM = "2026-05-23"
TERM_BREAK_TO = "2026-05-31"
TERM_RESUME_DATE = "2026-06-01"
TERM_CLOSED_DATES = ["2026-05-04"]
TERM_STAFF_AWAY_DATES_BY_PROFILE_KEY = {
    "roberto": ["2026-05-07"],
    "javier": ["2026-06-18", "2026-06-21"],
}
# Summer Term 2: staff no longer on pool rota these weekdays (term cell red).
TERM_STAFF_OFF_WEEKDAYS_RANGE_BY_PROFILE_KEY = {
    "roberto": {"from": "2026-06-01", "to": "2026-07-17", "weekdays": [6]},
}
# Term calendar green override when export/roster reconciliation misses a worked day.
TERM_STAFF_FEEDBACK_COMPLETE_DATES_BY_PROFILE_KEY = {
    "roberto": ["2026-05-17", "2026-05-22"],
    "javier": ["2026-05-12", "2026-05-13", "2026-05-14", "2026-05-17", "2026-05-19", "2026-05-20"],
    "youssef": [
        "2026-05-06",
        "2026-05-11",
        "2026-05-12",
        "2026-05-13",
        "2026-05-16",
        "2026-05-18",
        "2026-05-19",
        "2026-05-20",
    ],
}
# Staff may open specific pre-term calendar days for catch-up feedback (My Term grid).
TERM_STAFF_EXTRA_CALENDAR_DATES_BY_PROFILE_KEY = {
    "javier": [],
}
# Catch-up days: ignore static export greens; use real portal review state only.
TERM_STAFF_CATCH_UP_FEEDBACK_DATES_BY_PROFILE_KEY = {
    "javier": ["2026-05-14", "2026-05-17", "2026-05-20"],
}
# On catch-up days, treat these client slugs as already done (e.g. May 17 Samer done, Zaid pending).
TERM_STAFF_CATCH_UP_FEEDBACK_DONE_CLIENTS_BY_DATE_BY_PROFILE_KEY = {
    "javier": {
        "2026-05-17": [
            "shire",
            "samer",
            "hazem",
            "eiji",
            "rayyan_fi",
            "haneef",
            "max",
            "shaan",
        ],
        "2026-05-20": [
            "kayden",
        ],
    },
}
# Skip admin late-submission approval prompts (feedback/cancellation/incident).
TERM_STAFF_LATE_SUBMISSION_BYPASS_PROFILE_KEYS = ["javier"]
# First calendar date this client appears on rota (slug = clientId in staff bundle).
TERM_CLIENT_FIRST_SESSION_DATE = {
    "amaar_ah": "2026-04-24",
}


def norm_text(v):
    if v is None:
        return ""
    s = str(v).replace("\t", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_client_info_blob(v):
    """Preserve line breaks in numbered questionnaire blobs (General Info sheet)."""
    if v is None:
        return ""
    s = str(v).replace("\t", " ").replace("\r\n", "\n").replace("\r", "\n").strip()
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r" *\n *", "\n", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s


def to_iso_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = norm_text(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            d = datetime.strptime(s, fmt).date()
            if d.year < 2000:
                return None
            return d.isoformat()
        except ValueError:
            pass
    return None


def parse_staff_assignment(raw):
    t = norm_text(raw)
    if not t or t.lower() in {"-", "n/a", "na", "none"}:
        return None
    m = re.match(r"^(.*?)(\d[\d\.:]*\s*[-]\s*\d[\d\.:]*)$", t)
    if m:
        name = normalize_staff_display_name(norm_text(m.group(1)))
        timerange = norm_text(m.group(2)).replace(" ", "")
        return {"staff_name": name, "time_range": timerange, "raw": t}
    return {"staff_name": normalize_staff_display_name(t), "time_range": "", "raw": t}


def export_json_csv(base_name, rows):
    json_path = OUT / f"{base_name}.json"
    csv_path = OUT / f"{base_name}.csv"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=True, indent=2)

    fieldnames = sorted({k for row in rows for k in row.keys()}) if rows else []
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        if fieldnames:
            w.writeheader()
            for r in rows:
                w.writerow(r)


def patch_staff_timetable_swimfarm():
    path = SPREAD / "Staff Timetable (PORTAL).xlsx"
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Timetable"]
    sample = " ".join(
        str(ws.cell(row=r, column=c).value or "")
        for r in range(6, 12)
        for c in range(8, 11)
    )
    if "John 4.15-6.15" not in sample and "Bismark 4.15-6.15" not in sample:
        wb.close()
        return
    max_r = ws.max_row or 500
    for r in range(1, max_r + 1):
        date_iso = to_iso_date(ws.cell(row=r, column=2).value)
        for c in range(1, 18):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if not isinstance(v, str):
                continue
            if date_iso == "2026-05-04" and c in (8, 9, 10):
                cell.value = "n/a"
                continue
            new_v = (
                v.replace("John 4.15-6.15", "Victor 4.15-6.15")
                .replace("Bismark 4.15-6.15", "Javi 4.15-6.15")
                .replace("Giuseppe 4.15-6.15", "Raul 4.15-6.15")
            )
            if new_v != v:
                cell.value = new_v
    wb.save(path)


def portal_staff_profile_key(staff_name: str) -> str:
    """First token, lower, strip accents; align keys with staff_profiles / spreadsheet bundle."""
    t = norm_text(staff_name)
    if not t:
        return ""
    first = (t.split()[0] if t.split() else t).strip()
    nf = unicodedata.normalize("NFD", first)
    ascii_first = "".join(c for c in nf if unicodedata.category(c) != "Mn")
    v = re.sub(r"[^a-z0-9]+", "", ascii_first.lower())
    if v in ("yousef", "youssef", "yusef"):
        return "youssef"
    return v


ROSTER_CLIENT_CANONICAL = "Yusuf Ah"
ROSTER_CLIENT_ALIASES = frozenset({"Yusuf", "Yusef", "Yusuf Ah", "Yusuf Ahmed"})
STAFF_DISPLAY_YOUSSEF = "Youssef"


def normalize_roster_client_name(name: str) -> str:
    t = norm_text(name)
    return ROSTER_CLIENT_CANONICAL if t in ROSTER_CLIENT_ALIASES else t


def normalize_staff_display_name(name: str) -> str:
    t = norm_text(name)
    if t.lower() in ("yusef", "yousef", "youssef"):
        return STAFF_DISPLAY_YOUSSEF
    return t


def term_staff_weekday_indices_from_timetable_records(records: list) -> dict:
    """
    Weekday indices for TERM grid (same numbering as portalWorkedWeekdaysFromSessions):
    Sunday=0 .. Saturday=6. Built from Staff Timetable (PORTAL).xlsx Timetable sheet rows.
    """
    day_to_grid = {
        "monday": 1,
        "tuesday": 2,
        "wednesday": 3,
        "thursday": 4,
        "friday": 5,
        "saturday": 6,
        "sunday": 0,
    }
    by_key: dict[str, set[int]] = {}
    for r in records or []:
        pk = portal_staff_profile_key(str(r.get("staff_name") or ""))
        if not pk:
            continue
        dkey = norm_text(r.get("day") or "").lower()
        if dkey not in day_to_grid:
            continue
        by_key.setdefault(pk, set()).add(day_to_grid[dkey])
    return {k: sorted(v) for k, v in sorted(by_key.items()) if v}


def term_staff_weekday_indices_from_roster_machine_rows(rows: list) -> dict:
    """
    Same Sun=0..Sat=6 indices from flat roster rows (staff_clients_machine.json / xlsx).
    Merges with timetable-derived weekdays so staff who work Sundays on the roster
    still get TERM blues when the pool timetable grid omits that day.
    """
    day_to_grid = {
        "monday": 1,
        "tuesday": 2,
        "wednesday": 3,
        "thursday": 4,
        "friday": 5,
        "saturday": 6,
        "sunday": 0,
    }
    by_key: dict[str, set[int]] = {}
    for r in rows or []:
        day_raw = norm_text(r.get("day") or "").lower()
        if day_raw not in day_to_grid:
            continue
        wi = day_to_grid[day_raw]
        inst_blob = norm_text(r.get("instructors") or "")
        if not inst_blob:
            continue
        for piece in re.split(r"[,/&]+|\s+and\s+", inst_blob, flags=re.IGNORECASE):
            pk = portal_staff_profile_key(piece.strip())
            if not pk:
                continue
            by_key.setdefault(pk, set()).add(wi)
    return {k: sorted(v) for k, v in sorted(by_key.items()) if v}


def merge_term_staff_weekday_maps(tt: dict, roster: dict) -> dict:
    keys = set(tt) | set(roster)
    out = {}
    for k in sorted(keys):
        s = set(tt.get(k) or []) | set(roster.get(k) or [])
        if s:
            out[k] = sorted(s)
    return out


def merge_dashboard_weekday_maps(tt: dict, roster: dict) -> dict:
    """
    Timetable weekdays for the dashboard term grid, plus roster-only Mon–Sat days
    (e.g. Victor Thu Bespoke). Roster Sunday (0) is excluded so pool-only Sunday
    slots do not turn the whole term grid blue for swimming staff.
    """
    out: dict[str, list[int]] = {}
    keys = set(tt) | set(roster)
    for k in sorted(keys):
        s = set(tt.get(k) or [])
        for d in roster.get(k) or []:
            if int(d) != 0:
                s.add(int(d))
        if s:
            out[k] = sorted(s)
    return out


def term_staff_shift_dates_from_roster_machine_rows(
    rows: list, from_iso: str, to_iso: str
) -> dict:
    """Per-staff calendar dates from flat roster (e.g. CEO Thu Bespoke not on pool timetable)."""
    out: dict[str, set[str]] = {}
    f = str(from_iso or "")[:10]
    t = str(to_iso or "")[:10]
    for r in rows or []:
        iso = str(r.get("session_date") or "")[:10]
        if not iso or len(iso) < 10:
            continue
        if f and iso < f:
            continue
        if t and iso > t:
            continue
        inst_blob = norm_text(r.get("instructors") or "")
        if not inst_blob:
            continue
        for piece in re.split(r"[,/&]+|\s+and\s+", inst_blob, flags=re.IGNORECASE):
            pk = portal_staff_profile_key(piece.strip())
            if not pk:
                continue
            out.setdefault(pk, set()).add(iso)
    return {k: sorted(v) for k, v in sorted(out.items()) if v}


def merge_term_staff_shift_date_maps(tt: dict, roster: dict) -> dict:
    keys = set(tt) | set(roster)
    out = {}
    for k in sorted(keys):
        merged = sorted(set(tt.get(k) or []) | set(roster.get(k) or []))
        if merged:
            out[k] = merged
    return out


def term_staff_shift_dates_by_profile_key(
    records: list, from_iso: str, to_iso: str
) -> dict:
    """Per-staff calendar dates with a pool shift (Staff Timetable), inclusive range."""
    out: dict[str, set[str]] = {}
    f = str(from_iso or "")[:10]
    t = str(to_iso or "")[:10]
    for r in records or []:
        iso = str(r.get("date") or "")[:10]
        if not iso or len(iso) < 10:
            continue
        if f and iso < f:
            continue
        if t and iso > t:
            continue
        pk = portal_staff_profile_key(str(r.get("staff_name") or ""))
        if not pk:
            continue
        out.setdefault(pk, set()).add(iso)
    return {k: sorted(v) for k, v in sorted(out.items()) if v}


def _month_range_keys(from_iso: str, to_iso: str) -> list[tuple[int, int]]:
    fy, fm, _ = [int(x) for x in from_iso.split("-")]
    ly, lm, _ = [int(x) for x in to_iso.split("-")]
    keys: list[tuple[int, int]] = []
    y, m = fy, fm
    while (y, m) <= (ly, lm):
        keys.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return keys


def write_term_from_timetable_js(records, roster_rows=None):
    """JS snippet for dashboards: term grid date range + per-staff weekdays from Staff Timetable (PORTAL).xlsx."""
    path = OUT / "term_from_timetable.js"

    def _valid_term_date(s):
        if not s or len(s) < 10:
            return False
        try:
            return int(s[:4]) >= 2020
        except ValueError:
            return False

    dates = sorted({r["date"] for r in records if r.get("date") and _valid_term_date(r["date"])})
    if not dates:
        return
    first_s, last_s = dates[0], dates[-1]
    fy, fm, _ = [int(x) for x in first_s.split("-")]
    ly, lm, _ = [int(x) for x in last_s.split("-")]
    month_keys = []
    y, m = fy, fm
    while (y, m) <= (ly, lm):
        month_keys.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    term_calendar_months = [mm - 1 for _, mm in month_keys]
    term_calendar_year = fy
    first_dom_map = {}
    for mi in term_calendar_months:
        prefix = f"{term_calendar_year:04d}-{mi + 1:02d}-"
        in_month = [d for d in dates if d.startswith(prefix)]
        if not in_month:
            continue
        first_day = min(int(d.split("-")[2]) for d in in_month)
        if first_day > 1:
            first_dom_map[str(mi)] = first_day
    # Monday ISO dates starting a half-term week (timetable may omit; adjust when sourced from sheet).
    half_term_week_starts = ["2026-05-25"]
    wd_tt = term_staff_weekday_indices_from_timetable_records(records)
    wd_roster = term_staff_weekday_indices_from_roster_machine_rows(roster_rows or [])
    staff_wd = merge_term_staff_weekday_maps(wd_tt, wd_roster)
    view_from = TERM_RESUME_DATE
    view_to = last_s if last_s >= view_from else TERM_RESUME_DATE
    tt_summer = [
        r
        for r in records
        if r.get("date") and str(r["date"])[:10] >= view_from
    ]
    roster_summer = [
        r
        for r in roster_rows or []
        if not norm_text(r.get("session_date"))
        or str(r.get("session_date") or "")[:10] >= view_from
    ]
    wd_tt_summer = term_staff_weekday_indices_from_timetable_records(tt_summer)
    wd_roster_summer = term_staff_weekday_indices_from_roster_machine_rows(roster_summer)
    # Dashboard pattern weekdays: timetable + roster Mon–Sat (Thu bespoke for Victor, etc.).
    staff_wd_dashboard = merge_dashboard_weekday_maps(wd_tt_summer, wd_roster_summer)
    shift_dates_tt = term_staff_shift_dates_by_profile_key(
        tt_summer, view_from, view_to
    )
    shift_dates_roster = term_staff_shift_dates_from_roster_machine_rows(
        roster_summer, view_from, view_to
    )
    shift_dates_dashboard = merge_term_staff_shift_date_maps(
        shift_dates_tt, shift_dates_roster
    )
    view_month_keys = _month_range_keys(view_from, view_to)
    dashboard_months = [mm - 1 for _, mm in view_month_keys]
    dashboard_year = view_month_keys[0][0] if view_month_keys else term_calendar_year
    dashboard_first_dom: dict[str, int] = {}
    for mi in dashboard_months:
        prefix = f"{dashboard_year:04d}-{mi + 1:02d}-"
        in_month = [
            d
            for d in dates
            if d.startswith(prefix) and view_from <= d <= view_to
        ]
        if not in_month:
            continue
        first_day = min(int(d.split("-")[2]) for d in in_month)
        if first_day > 1:
            dashboard_first_dom[str(mi)] = first_day
    payload = {
        "termName": "Summer Term 2026",
        "termCalendarYear": term_calendar_year,
        "termCalendarMonths": term_calendar_months,
        "termCalendarFirstDom": first_dom_map,
        "termDashboardCalendarYear": dashboard_year,
        "termDashboardCalendarMonths": dashboard_months,
        "termDashboardCalendarFirstDom": dashboard_first_dom,
        "termDashboardCalendarFrom": view_from,
        "termDashboardCalendarTo": view_to,
        "firstDate": first_s,
        "lastDate": last_s,
        "termBreakFrom": TERM_BREAK_FROM,
        "termBreakTo": TERM_BREAK_TO,
        "termResumeDate": TERM_RESUME_DATE,
        "termClosedDates": list(TERM_CLOSED_DATES),
        "termStaffAwayDatesByProfileKey": dict(TERM_STAFF_AWAY_DATES_BY_PROFILE_KEY),
        "termStaffOffWeekdaysRangeByProfileKey": dict(
            TERM_STAFF_OFF_WEEKDAYS_RANGE_BY_PROFILE_KEY
        ),
        "termStaffFeedbackCompleteDatesByProfileKey": dict(
            TERM_STAFF_FEEDBACK_COMPLETE_DATES_BY_PROFILE_KEY
        ),
        "termStaffExtraCalendarDatesByProfileKey": dict(
            TERM_STAFF_EXTRA_CALENDAR_DATES_BY_PROFILE_KEY
        ),
        "termStaffCatchUpFeedbackDatesByProfileKey": dict(
            TERM_STAFF_CATCH_UP_FEEDBACK_DATES_BY_PROFILE_KEY
        ),
        "termStaffCatchUpFeedbackDoneClientsByDateByProfileKey": dict(
            TERM_STAFF_CATCH_UP_FEEDBACK_DONE_CLIENTS_BY_DATE_BY_PROFILE_KEY
        ),
        "termStaffLateSubmissionBypassProfileKeys": list(
            TERM_STAFF_LATE_SUBMISSION_BYPASS_PROFILE_KEYS
        ),
        "termClientFirstSessionDate": dict(TERM_CLIENT_FIRST_SESSION_DATE),
        "termHalfTermWeekStarts": half_term_week_starts,
        "termStaffWeekdayIndicesByProfileKey": staff_wd,
        "termStaffWeekdayIndicesDashboardByProfileKey": staff_wd_dashboard,
        "termStaffShiftDatesByProfileKey": shift_dates_dashboard,
    }
    body = (
        "// Auto-generated by database/build_machine_exports.py from Staff Timetable (PORTAL).xlsx\n"
        "// Re-run: python database/build_machine_exports.py\n"
        "// termStaffWeekdayIndicesByProfileKey = weekdays (Sun=0..Sat=6) from timetable + staff_clients_machine roster.\n"
        "// termStaffShiftDatesByProfileKey = exact shift dates (timetable + roster machine) for My Term blue/red from 1 Jun.\n"
        "window.PORTAL_TERM_FROM_TIMETABLE = "
        + json.dumps(payload, indent=2)
        + ";\n"
    )
    path.write_text(body, encoding="utf-8")


def build_staff_timetable():
    wb = openpyxl.load_workbook(SPREAD / "Staff Timetable (PORTAL).xlsx", data_only=True)
    ws = wb["Timetable"]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    current_day = None
    i = 0
    while i < len(rows):
        r = rows[i]
        c2 = norm_text(r[1] if len(r) > 1 else None)
        if c2.lower() in {
            "mondays", "tuesdays", "wednesdays", "thursdays", "fridays", "saturdays", "sundays"
        }:
            current_day = c2.rstrip("s")
            if i + 1 < len(rows):
                headers = rows[i + 1]
                if norm_text(headers[1] if len(headers) > 1 else None).lower() == "dates":
                    venues = {}
                    for col in range(2, len(headers)):
                        h = norm_text(headers[col])
                        if h:
                            venues[col] = h
                    j = i + 2
                    while j < len(rows):
                        rr = rows[j]
                        first = norm_text(rr[1] if len(rr) > 1 else None)
                        if not first:
                            j += 1
                            continue
                        if first.lower() in {
                            "mondays", "tuesdays", "wednesdays", "thursdays", "fridays", "saturdays", "sundays"
                        }:
                            break
                        d = to_iso_date(rr[1] if len(rr) > 1 else None)
                        if not d:
                            j += 1
                            continue
                        for col, venue in venues.items():
                            raw = rr[col] if col < len(rr) else None
                            parsed = parse_staff_assignment(raw)
                            if not parsed:
                                continue
                            records.append({
                                "day": current_day,
                                "date": d,
                                "venue": venue,
                                "staff_name": normalize_staff_display_name(parsed["staff_name"]),
                                "time_range": parsed["time_range"],
                                "raw_assignment": parsed["raw"],
                            })
                        j += 1
                    i = j
                    continue
        i += 1

    export_json_csv("staff_timetable_machine", records)
    return records


def read_staff_clients_flat_workbook(path: Path) -> list:
    """Machine-export sheet: client_name, day, instructors, service, area, time_slot, venue (header row)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [norm_text(c).lower().replace(" ", "_") for c in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    def col(*names):
        for n in names:
            if n in idx:
                return idx[n]
        return None

    ci = col("client_name", "client")
    di = col("day")
    ii = col("instructors", "instructor", "staff")
    si = col("service", "programme", "program")
    ai = col("area", "pool", "zone")
    ti = col("time_slot", "time", "slot")
    vi = col("venue", "location")
    out = []
    for r in rows[1:]:
        def get(i):
            if i is None or i >= len(r):
                return ""
            return norm_text(r[i])

        rec = {
            "client_name": normalize_roster_client_name(get(ci)),
            "day": get(di),
            "instructors": get(ii),
            "service": get(si),
            "area": get(ai),
            "time_slot": get(ti),
            "venue": get(vi),
        }
        if not any(rec.values()):
            continue
        out.append(rec)
    return out


def build_staff_clients():
    """Roster rows: only `database/staff_clients_machine.xlsx` (flat machine export)."""
    flat_path = OUT / "staff_clients_machine.xlsx"
    if not flat_path.exists():
        raise FileNotFoundError(
            "Missing "
            + str(flat_path)
            + " - add the flat workbook (columns: client_name, day, instructors, service, area, time_slot, venue)."
        )
    records = read_staff_clients_flat_workbook(flat_path)
    export_json_csv("staff_clients_machine", records)


def build_clients_info():
    candidates = (
        ROOT / "working_ui" / "PARTICIPANTS. GENERAL INFO.xlsx",
        SPREAD / "PARTICIPANTS. GENERAL INFO.xlsx",
        SPREAD / "Clients Info (PORTAL).xlsx",
        ROOT / "working_ui" / "Clients Info (PORTAL).xlsx",
    )
    path = next((p for p in candidates if p.exists()), None)
    if not path:
        print("build_clients_info: no workbook found (PARTICIPANTS. GENERAL INFO.xlsx or Clients Info)")
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = next(
        (sn for sn in ("General info", "General Info", "Clients info", "Clients Info") if sn in wb.sheetnames),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = norm_text(r[0] if len(r) > 0 else None)
        info = norm_client_info_blob(r[1] if len(r) > 1 else None)
        if not name or not info:
            continue
        rows.append({"client_name": name, "client_info": info})
    export_json_csv("clients_info_machine", rows)
    print(f"build_clients_info: {len(rows)} rows from {path.name} ({sheet_name})")


def write_clients_info_embed_js():
    path_json = OUT / "clients_info_machine.json"
    if not path_json.exists():
        return
    rows = json.loads(path_json.read_text(encoding="utf-8"))
    body = (
        "// Auto-generated by database/build_machine_exports.py\n"
        "window.PORTAL_CLIENTS_INFO_ROWS = "
        + json.dumps(rows, ensure_ascii=True)
        + ";\n"
    )
    targets = (
        OUT / "clients_info_embed.js",
        ROOT / "working_ui" / "clients_info_embed.js",
        ROOT / "working_ui" / "portal" / "clients_info_embed.js",
        ROOT / "working_ui" / "portal-shared-js" / "clients_info_embed.js",
    )
    for dst in targets:
        dst.parent.mkdir(parents=True, exist_ok=True)
        dst.write_text(body, encoding="utf-8")


def copy_term_to_working_ui():
    src = OUT / "term_from_timetable.js"
    if not src.exists():
        return
    text = src.read_text(encoding="utf-8")
    for rel in (
        "term_from_timetable.js",
        "portal/term_from_timetable.js",
        "portal-shared-js/term_from_timetable.js",
    ):
        (ROOT / "working_ui" / rel).write_text(text, encoding="utf-8")


def copy_spreadsheet_js_to_working_ui():
    for name in ("staff_dashboard_spreadsheet_adapter.js", "staff_dashboard_spreadsheet_bundle.js"):
        src = OUT / name
        if not src.exists():
            continue
        text = normalize_bundle_portal_asset_urls(src.read_text(encoding="utf-8"))
        (ROOT / "working_ui" / name).write_text(text, encoding="utf-8")
        (ROOT / "working_ui" / "portal" / name).write_text(text, encoding="utf-8")


def patch_bundle_rows_from_json():
    """Keep embedded bundle rows in sync with staff_clients_machine.json."""
    bundle_path = OUT / "staff_dashboard_spreadsheet_bundle.js"
    json_path = OUT / "staff_clients_machine.json"
    if not bundle_path.exists() or not json_path.exists():
        return
    rows = json.loads(json_path.read_text(encoding="utf-8"))
    src = bundle_path.read_text(encoding="utf-8")
    needle = '"rows":'
    i = src.find(needle)
    if i < 0:
        return
    j = src.find("[", i)
    if j < 0:
        return
    depth = 0
    k = j
    end = -1
    while k < len(src):
        c = src[k]
        if c == "[":
            depth += 1
        elif c == "]":
            depth -= 1
            if depth == 0:
                end = k + 1
                break
        k += 1
    if end < 0:
        return
    new_rows = json.dumps(rows, ensure_ascii=True, indent=2)
    new_src = src[:j] + new_rows + src[end:]
    new_src = normalize_bundle_portal_asset_urls(new_src)
    bundle_path.write_text(new_src, encoding="utf-8")


def normalize_bundle_portal_asset_urls(text: str) -> str:
    """Same-origin relative paths for Vercel/static deploy (not clubsensational.org CDN)."""
    return text.replace(
        "https://www.clubsensational.org/portal/staff_photos/",
        "portal/staff_photos/",
    )


if __name__ == "__main__":
    patch_staff_timetable_swimfarm()
    timetable_records = build_staff_timetable()
    build_staff_clients()
    try:
        from import_roster_week_csv import import_all_roster_weeks

        n_week = import_all_roster_weeks()
        if n_week:
            print(f"Merged {n_week} dated rows from database/roster_weeks/ (+ portal week CSV)")
    except Exception as exc:
        print("import_roster_week_csv:", exc)
    try:
        from import_notas_participants_portal import sync_notas

        notas_stats = sync_notas()
        if notas_stats:
            print(
                f"NOTAS participants: {notas_stats['parsed']} slots, "
                f"{notas_stats['updated']} roster rows updated"
            )
    except Exception as exc:
        print("import_notas_participants_portal:", exc)
    roster_path = OUT / "staff_clients_machine.json"
    roster_rows = (
        json.loads(roster_path.read_text(encoding="utf-8"))
        if roster_path.exists()
        else []
    )
    write_term_from_timetable_js(timetable_records, roster_rows)
    build_clients_info()
    write_clients_info_embed_js()
    patch_bundle_rows_from_json()
    try:
        from patch_teflon_demo_environment import patch_machine_json, patch_bundle_file

        patch_machine_json()
        patch_bundle_file(OUT / "staff_dashboard_spreadsheet_bundle.js")
    except Exception as exc:
        print("patch_teflon_demo_environment:", exc)
    try:
        from patch_roster_bundle_feedback_rules import patch_all as patch_feedback_rules

        patch_feedback_rules()
    except Exception as exc:
        print("patch_roster_bundle_feedback_rules:", exc)
    copy_term_to_working_ui()
    copy_spreadsheet_js_to_working_ui()
    print("Generated machine files in database/ and term_from_timetable.js")

