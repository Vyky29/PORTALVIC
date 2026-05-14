"""
Build a machine-readable workbook from Clients Info free-text blobs.

Source: database/clients_info_machine.json (from Clients Info (PORTAL).xlsx via build_clients_info).

Output: working_ui/Clients Info (PORTAL) - structured.xlsx
  Row 1: fixed headers (Participant Name + 15 assessment fields).
  Data: one row per client; cells filled from parsed 1. Age ... 15. Other Notes sections.

Run: python database/build_clients_info_structured_xlsx.py
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SPREAD = ROOT / "SPREADSHEETS"
LEGACY_XLSX = SPREAD / "Clients Info (PORTAL).xlsx"
JSON_PATH = ROOT / "database" / "clients_info_machine.json"
OUT_XLSX = ROOT / "working_ui" / "Clients Info (PORTAL) - structured.xlsx"


def load_rows_from_legacy_xlsx() -> list[dict] | None:
    if not LEGACY_XLSX.exists():
        return None
    wb = openpyxl.load_workbook(LEGACY_XLSX, data_only=True)
    if "Clients info" not in wb.sheetnames:
        return None
    ws = wb["Clients info"]
    out: list[dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = norm_text(r[0] if len(r) > 0 else None)
        info = norm_text(r[1] if len(r) > 1 else None)
        if not name:
            continue
        out.append({"client_name": name, "client_info": info})
    return out if out else None

# Headers match operator numbering 0..15
HEADERS = [
    "Participant Name",
    "Age",
    "Medical",
    "Likes/Motivators",
    "Dislikes/Avoids",
    "Known Triggers",
    "Regulation Strategies",
    "Level of Support",
    "Communication",
    "Preferred Communication",
    "Mobility",
    "Personal Care",
    "Task Engagement",
    "Transitions/Flexibility",
    "Safety",
    "Other Notes",
]

# Strip leading label after "N. " for each section number (case-insensitive).
_STRIP = {
    1: re.compile(r"^Age:\s*", re.I),
    2: re.compile(r"^Medical:\s*", re.I),
    3: re.compile(r"^Likes/Motivators:\s*", re.I),
    4: re.compile(r"^Dislikes/Avoids:\s*", re.I),
    5: re.compile(r"^Known Triggers:\s*", re.I),
    6: re.compile(r"^Regulation Strategies:\s*", re.I),
    7: re.compile(r"^Level of Support:\s*", re.I),
    8: re.compile(r"^Communication:\s*", re.I),
    9: re.compile(r"^Preferred Communication:\s*", re.I),
    10: re.compile(r"^Mobility:\s*", re.I),
    11: re.compile(r"^Personal Care:\s*", re.I),
    12: re.compile(r"^Task Engagement:\s*", re.I),
    13: re.compile(r"^Transitions/Flexibility:\s*", re.I),
    14: re.compile(r"^Safety:\s*", re.I),
    15: re.compile(r"^Other Notes:\s*", re.I),
}


def norm_text(v: object) -> str:
    if v is None:
        return ""
    s = str(v).replace("\t", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def parse_numbered_sections(blob: str) -> dict[int, str]:
    """Split '1. Age: ... 2. Medical: ...' into {1: 'Age: ...', 2: 'Medical: ...', ...}."""
    text = (blob or "").strip()
    if not text:
        return {}
    matches = list(re.finditer(r"(?:^|\s)(\d{1,2})\.\s+", text))
    out: dict[int, str] = {}
    for i, m in enumerate(matches):
        num = int(m.group(1))
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        chunk = text[start:end].strip()
        if num in _STRIP:
            chunk = _STRIP[num].sub("", chunk, count=1).strip()
        out[num] = chunk
    return out


def row_from_record(rec: dict) -> list[str]:
    name = norm_text(rec.get("client_name"))
    blob = norm_text(rec.get("client_info"))
    sections = parse_numbered_sections(blob)
    row = [name]
    for n in range(1, 16):
        row.append(sections.get(n, ""))
    return row


def main() -> None:
    rows_data = load_rows_from_legacy_xlsx()
    source = str(LEGACY_XLSX)
    if rows_data is None:
        if not JSON_PATH.exists():
            raise SystemExit(
                f"Missing {LEGACY_XLSX} and {JSON_PATH} — add the source workbook under SPREADSHEETS/ "
                "or run build_machine_exports.py to create clients_info_machine.json."
            )
        rows_data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
        source = str(JSON_PATH)
    if not isinstance(rows_data, list):
        raise SystemExit("clients_info rows must be a list of objects with client_name and client_info")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients info"

    header_font = Font(bold=True)
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.alignment = Alignment(vertical="top", wrap_text=True)

    for ridx, rec in enumerate(rows_data, start=2):
        if not isinstance(rec, dict):
            continue
        for cidx, val in enumerate(row_from_record(rec), start=1):
            cell = ws.cell(row=ridx, column=cidx, value=val or None)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Sensible widths (Participant + long text columns)
    widths = [22, 28, 40, 48, 28, 32, 36, 32, 32, 36, 28, 32, 36, 32, 36, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 60)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX} ({len(rows_data)} data rows + header) from {source}.")


if __name__ == "__main__":
    main()
