# -*- coding: utf-8 -*-
"""
Parse working_ui/NOTAS participants (PORTAL).xlsx and sync area (notes), venue,
and time_slot onto database/staff_clients_machine.json (+ xlsx + spreadsheet bundle).

The NOTAS sheet is a weekday grid: venue columns × (time | participant | notes).
Instructors and service stay from the existing machine roster; only pool/room notes
and anchor time/venue are updated from NOTAS.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
NOTAS_PATH = ROOT / "working_ui" / "NOTAS participants (PORTAL).xlsx"
MACHINE_JSON = ROOT / "database" / "staff_clients_machine.json"
MACHINE_XLSX = ROOT / "database" / "staff_clients_machine.xlsx"
BUNDLE_JS = ROOT / "database" / "staff_dashboard_spreadsheet_bundle.js"

DAYS = frozenset(
    {"SATURDAY", "SUNDAY", "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"}
)
SKIP_CLIENTS = frozenset(
    {"CLOSED", "NO PARTICIPANT", "NO CLIENT", "N/A", "NA", "NONE", "-", ""}
)
ROSTER_CLIENT_CANONICAL = "Yusuf Ah"
ROSTER_CLIENT_ALIASES = frozenset({"Yusuf", "Yusef", "Yusuf Ah", "Yusuf Ahmed"})


def norm_text(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\n", " ").replace("\t", " ")
    return re.sub(r"\s+", " ", s).strip()


def normalize_client_name(name: str) -> str:
    t = norm_text(name)
    return ROSTER_CLIENT_CANONICAL if t in ROSTER_CLIENT_ALIASES else t


def normalize_time_slot(raw: str) -> str:
    t = norm_text(raw).lower()
    if not t:
        return ""
    t = t.replace(".", ":")
    t = re.sub(r"\s*-\s*", " to ", t)
    t = re.sub(r"\s+to\s+", " to ", t)
    t = re.sub(r"(\d)(to)", r"\1 \2", t)
    t = re.sub(r"(?<=\d)\s+(?=\d)", ":", t, count=1) if False else t
    # "4.30-5.30" / "4.30 to 5"
    t = re.sub(r"(\d{1,2}):(\d{2})to", r"\1:\2 to ", t)
    t = re.sub(r"(\d{1,2}) to (\d{1,2}):(\d{2})", r"\1 to \2:\3", t)
    t = re.sub(r"(\d{1,2}):(\d{2}) to (\d{1,2})\b(?!:)", r"\1:\2 to \3", t)
    t = re.sub(r"(\d{1,2}) to (\d{1,2})\b(?!:)", r"\1 to \2", t)
    return norm_text(t)


def time_match_key(raw: str) -> str:
    t = normalize_time_slot(raw)
    t = t.replace(":", ".")
    return re.sub(r"\s+", "", t.lower())


def client_match_key(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_client_name(name).lower())


def venue_match_key(v: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", norm_text(v).lower())


def row_cells(ws, r: int) -> dict[int, str]:
    out: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        v = norm_text(ws.cell(r, c).value)
        if v:
            out[c] = v
    return out


def venue_for_col(venue_marks: list[tuple[int, str]], col: int) -> str:
    best = ""
    for vc, vn in venue_marks:
        if vc <= col:
            best = vn
    return best


def parse_notas_workbook(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records: list[dict] = []
    current_day: str | None = None
    blocks: list[dict] = []
    block_times: dict[int, str] = {}

    for r in range(1, ws.max_row + 1):
        cells = row_cells(ws, r)
        if not cells:
            continue
        b = cells.get(2, "")

        if b.upper() in DAYS:
            current_day = b.strip().title()
            blocks = []
            block_times = {}
            continue
        if not current_day:
            continue

        # Header row: TIME | PARTICIPANTS | NOTES (repeated per venue block)
        time_cols = sorted(c for c, v in cells.items() if v.upper() == "TIME")
        if time_cols:
            prev = row_cells(ws, r - 1) if r > 1 else {}
            venue_marks = sorted(
                (c, prev[c])
                for c in prev
                if prev[c].upper() not in DAYS
                and prev[c].upper() not in {"TIME", "PARTICIPANTS", "PARTICIPANT", "NOTES", "POOL"}
            )
            blocks = []
            for tc in time_cols:
                pc, nc = tc + 1, tc + 2
                note_hdr = cells.get(nc, "").upper()
                if note_hdr in {"NOTES", "POOL"}:
                    blocks.append(
                        {
                            "time_col": tc,
                            "part_col": pc,
                            "note_col": nc,
                            "venue": venue_for_col(venue_marks, tc),
                            "time_optional": False,
                        }
                    )
            # Westway wall: PARTICIPANTS | NOTES without TIME (time from sibling block)
            for c, v in cells.items():
                if v.upper() not in {"PARTICIPANTS", "PARTICIPANT"}:
                    continue
                if any(blk["part_col"] == c for blk in blocks):
                    continue
                nc = c + 1
                if cells.get(nc, "").upper() not in {"NOTES", "POOL"}:
                    continue
                blocks.append(
                    {
                        "time_col": None,
                        "part_col": c,
                        "note_col": nc,
                        "venue": venue_for_col(venue_marks, c),
                        "time_optional": True,
                    }
                )
            block_times = {}
            continue

        if not blocks:
            continue

        row_time_fallback = ""
        for blk in blocks:
            if blk["time_col"] and cells.get(blk["time_col"]):
                row_time_fallback = cells[blk["time_col"]]

        for blk in blocks:
            tc = blk["time_col"]
            pc = blk["part_col"]
            nc = blk["note_col"]
            venue = norm_text(blk["venue"])
            participant = norm_text(cells.get(pc, ""))
            area = norm_text(cells.get(nc, ""))
            if not participant or participant.upper() in SKIP_CLIENTS:
                continue
            if not venue:
                continue

            time_slot = norm_text(cells.get(tc, "")) if tc else ""
            if time_slot:
                if tc is not None:
                    block_times[tc] = time_slot
            elif tc is not None:
                time_slot = block_times.get(tc, "")
            if not time_slot:
                time_slot = row_time_fallback
            if not time_slot and blk["time_optional"]:
                continue
            if not time_slot:
                continue

            records.append(
                {
                    "day": current_day,
                    "venue": venue,
                    "time_slot": time_slot,
                    "client_name": normalize_client_name(participant),
                    "area": area,
                }
            )

    return records


def notas_lookup(records: list[dict]) -> dict[tuple, dict]:
    """Key = day, venue, client, time — value includes area."""
    out: dict[tuple, dict] = {}
    for r in records:
        key = (
            norm_text(r["day"]).lower(),
            venue_match_key(r["venue"]),
            client_match_key(r["client_name"]),
            time_match_key(r["time_slot"]),
        )
        out[key] = r
    return out


def fuzzy_find_notas(
    machine_row: dict, by_exact: dict[tuple, dict], all_notas: list[dict]
) -> dict | None:
    day_k = norm_text(machine_row.get("day")).lower()
    ven_k = venue_match_key(machine_row.get("venue"))
    cli_k = client_match_key(machine_row.get("client_name"))
    time_k = time_match_key(machine_row.get("time_slot"))

    hit = by_exact.get((day_k, ven_k, cli_k, time_k))
    if hit:
        return hit

    # Same day + venue + client; time overlap (NOTAS vs roster formatting)
    candidates = [
        n
        for n in all_notas
        if norm_text(n["day"]).lower() == day_k
        and venue_match_key(n["venue"]) == ven_k
        and client_match_key(n["client_name"]) == cli_k
    ]
    if len(candidates) == 1:
        return candidates[0]
    if not candidates:
        return None
    # Prefer closest time key prefix
    for n in candidates:
        if time_match_key(n["time_slot"]) == time_k:
            return n
    for n in candidates:
        nk = time_match_key(n["time_slot"])
        if nk.startswith(time_k[:4]) or time_k.startswith(nk[:4]):
            return n
    return candidates[0] if len(candidates) == 1 else None


def apply_notas_to_machine(machine_rows: list[dict], notas_rows: list[dict]) -> tuple[list[dict], dict]:
    by_exact = notas_lookup(notas_rows)
    stats = {"updated": 0, "unchanged": 0, "no_match": 0, "fields": {"area": 0, "time_slot": 0, "venue": 0}}
    out = []
    for row in machine_rows:
        rec = dict(row)
        hit = fuzzy_find_notas(rec, by_exact, notas_rows)
        if not hit:
            stats["no_match"] += 1
            out.append(rec)
            continue
        changed = False
        new_area = norm_text(hit.get("area"))
        if new_area and norm_text(rec.get("area")) != new_area:
            rec["area"] = new_area
            stats["fields"]["area"] += 1
            changed = True
        new_time = norm_text(hit.get("time_slot"))
        if new_time and time_match_key(rec.get("time_slot")) != time_match_key(new_time):
            rec["time_slot"] = new_time
            stats["fields"]["time_slot"] += 1
            changed = True
        new_venue = norm_text(hit.get("venue"))
        if new_venue and venue_match_key(rec.get("venue")) != venue_match_key(new_venue):
            rec["venue"] = new_venue
            stats["fields"]["venue"] += 1
            changed = True
        if changed:
            stats["updated"] += 1
        else:
            stats["unchanged"] += 1
        out.append(rec)
    return out, stats


def write_machine_xlsx(rows: list[dict], path: Path) -> None:
    headers = ["client_name", "day", "instructors", "service", "area", "time_slot", "venue", "session_date"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "roster"
    ws.append(headers)
    for r in rows:
        ws.append([norm_text(r.get(h)) for h in headers])
    wb.save(path)


def patch_bundle_rows(rows: list[dict]) -> None:
    if not BUNDLE_JS.exists():
        return
    src = BUNDLE_JS.read_text(encoding="utf-8")
    needle = '"rows":'
    i = src.find(needle)
    if i < 0:
        return
    j = src.find("[", i)
    if j < 0:
        return
    depth = 0
    k = j
    end = -1
    while k < len(src):
        c = src[k]
        if c == "[":
            depth += 1
        elif c == "]":
            depth -= 1
            if depth == 0:
                end = k + 1
                break
        k += 1
    if end < 0:
        return
    new_rows = json.dumps(rows, ensure_ascii=True, indent=2)
    BUNDLE_JS.write_text(src[:j] + new_rows + src[end:], encoding="utf-8")


def copy_artifacts(rows: list[dict], notas_rows: list[dict]) -> None:
    bundle_text = BUNDLE_JS.read_text(encoding="utf-8") if BUNDLE_JS.exists() else ""
    for rel in (
        "staff_dashboard_spreadsheet_bundle.js",
        "portal/staff_dashboard_spreadsheet_bundle.js",
        "portal-shared-js/staff_dashboard_spreadsheet_bundle.js",
    ):
        dst = ROOT / "working_ui" / rel
        if dst.parent.exists() or rel.startswith("portal"):
            dst.parent.mkdir(parents=True, exist_ok=True)
            if bundle_text:
                dst.write_text(bundle_text, encoding="utf-8")

    json_text = json.dumps(notas_rows, ensure_ascii=True, indent=2)
    (ROOT / "database" / "notas_participants_portal_parsed.json").write_text(
        json_text + "\n", encoding="utf-8"
    )


def sync_notas() -> dict | None:
    """Apply NOTAS xlsx to machine roster + bundle. Returns stats or None if skipped."""
    if not NOTAS_PATH.exists() or not MACHINE_JSON.exists():
        return None
    notas_rows = parse_notas_workbook(NOTAS_PATH)
    machine_rows = json.loads(MACHINE_JSON.read_text(encoding="utf-8"))
    updated_rows, stats = apply_notas_to_machine(machine_rows, notas_rows)
    MACHINE_JSON.write_text(json.dumps(updated_rows, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    write_machine_xlsx(updated_rows, MACHINE_XLSX)
    patch_bundle_rows(updated_rows)
    copy_artifacts(updated_rows, notas_rows)
    return {"parsed": len(notas_rows), **stats}


def main() -> int:
    if not NOTAS_PATH.exists():
        print("Missing", NOTAS_PATH, file=sys.stderr)
        return 1
    if not MACHINE_JSON.exists():
        print("Missing", MACHINE_JSON, file=sys.stderr)
        return 1

    result = sync_notas()
    if not result:
        return 1
    print(f"Parsed {result['parsed']} NOTAS slot rows from {NOTAS_PATH.name}")
    print(
        "Machine roster: "
        f"{result['updated']} rows updated, "
        f"{result['unchanged']} already matched, "
        f"{result['no_match']} without NOTAS match"
    )
    print(
        "Field changes: "
        f"area={result['fields']['area']}, "
        f"time_slot={result['fields']['time_slot']}, "
        f"venue={result['fields']['venue']}"
    )

    # Copy NOTAS source into database for traceability (not deployed — working_ui only reference)
    print("Updated", MACHINE_JSON.name, MACHINE_XLSX.name, "and spreadsheet bundle copies under working_ui/")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
