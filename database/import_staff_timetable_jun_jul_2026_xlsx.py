#!/usr/bin/env python3
"""
Import Summer Term 2 staff pool hours (2026-06-01 .. 2026-07-17) from the Jun–Jul timetable xlsx
into database/staff_timetable_machine.json, then refresh spreadsheet_reference_data.js.

Default source: working_ui/induction-assets/Timetable 1 de junio a 17 julio.xlsx

Run:
  python database/import_staff_timetable_jun_jul_2026_xlsx.py
  python database/export_spreadsheet_reference_js.py
"""
from __future__ import annotations

import json
import re
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "database"
DEFAULT_XLSX = ROOT / "working_ui" / "induction-assets" / "Timetable 1 de junio a 17 julio.xlsx"
TERM2_START = "2026-06-01"
TERM2_END = "2026-07-17"
WEEKDAYS = {
    "mondays": "Monday",
    "tuesdays": "Tuesday",
    "wednesdays": "Wednesday",
    "thursdays": "Thursday",
    "fridays": "Friday",
    "saturdays": "Saturday",
    "sundays": "Sunday",
}


def norm_text(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\t", " ").strip()
    return re.sub(r"\s+", " ", s)


def to_iso_date(v) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = norm_text(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            d = datetime.strptime(s, fmt).date()
            if d.year < 2000:
                return None
            return d.isoformat()
        except ValueError:
            pass
    return None


def normalize_staff_display_name(name: str) -> str:
    t = norm_text(name)
    if not t:
        return ""
    low = t.lower()
    aliases = {
        "raul": "Raul",
        "raúl": "Raul",
        "javi": "Javi",
        "luliya": "Luliya",
        "godsway": "Godsway",
        "bismark": "Bismark",
        "giuseppe": "Giuseppe",
    }
    for k, v in aliases.items():
        if low == k:
            return v
    return t[0].upper() + t[1:] if len(t) > 1 else t.upper()


def normalize_venue(h: str) -> str:
    t = norm_text(h).replace("\t", " ")
    if not t:
        return ""
    low = t.lower()
    if "westway" in low:
        return "Westway"
    if "northolt" in low:
        return "Northolt"
    if "acton" in low:
        return "Acton"
    if "swimfarm" in low or "swim farm" in low:
        return "SwimFarm"
    return t


def parse_staff_assignment(raw) -> dict | None:
    t = norm_text(raw)
    if not t or t.lower() in {"-", "n/a", "na", "none"}:
        return None
    m = re.match(r"^(.*?)(\d[\d\.:]*\s*[-]\s*\d[\d\.:]*)$", t)
    if m:
        name = normalize_staff_display_name(norm_text(m.group(1)))
        timerange = norm_text(m.group(2)).replace(" ", "")
        return {"staff_name": name, "time_range": timerange, "raw": t}
    name = normalize_staff_display_name(t)
    return {"staff_name": name, "time_range": "", "raw": t}


def build_venue_map(headers: tuple) -> dict[int, str]:
    """Column index -> venue; forward-fill merged header cells."""
    venues: dict[int, str] = {}
    last = ""
    for col in range(2, len(headers)):
        h = normalize_venue(headers[col] if col < len(headers) else "")
        if h:
            last = h
        if last:
            venues[col] = last
    return venues


def parse_timetable_sheet(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    records: list[dict] = []
    current_day: str | None = None
    venues: dict[int, str] = {}
    i = 0
    while i < len(rows):
        r = rows[i]
        c1 = norm_text(r[1] if len(r) > 1 else None).lower()
        if c1 in WEEKDAYS:
            current_day = WEEKDAYS[c1]
            venues = {}
            i += 1
            continue
        if c1 == "dates":
            venues = build_venue_map(r)
            i += 1
            continue
        if not current_day or not venues:
            i += 1
            continue
        iso = to_iso_date(r[1] if len(r) > 1 else None)
        if not iso:
            i += 1
            continue
        if iso < TERM2_START or iso > TERM2_END:
            i += 1
            continue
        sort_index = 0
        for col in sorted(venues.keys()):
            raw = r[col] if col < len(r) else None
            parsed = parse_staff_assignment(raw)
            if not parsed:
                continue
            records.append(
                {
                    "date": iso,
                    "day": current_day,
                    "venue": venues[col],
                    "staff_name": parsed["staff_name"],
                    "time_range": parsed["time_range"],
                    "raw_assignment": parsed["raw"],
                    "sort_index": sort_index,
                }
            )
            sort_index += 1
        i += 1
    return records


def merge_into_machine(new_rows: list[dict]) -> list[dict]:
    path = OUT / "staff_timetable_machine.json"
    existing: list[dict] = []
    if path.exists():
        existing = json.loads(path.read_text(encoding="utf-8"))
    kept = [r for r in existing if not (TERM2_START <= str(r.get("date") or "") <= TERM2_END)]
    merged = kept + new_rows
    merged.sort(
        key=lambda r: (
            str(r.get("date") or ""),
            str(r.get("day") or ""),
            int(r.get("sort_index") or 0),
            str(r.get("venue") or ""),
            str(r.get("raw_assignment") or ""),
        )
    )
    path.write_text(json.dumps(merged, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    return merged


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        raise SystemExit(f"Missing {xlsx}")

    import openpyxl

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    sheet_name = "Copy of Timetable" if "Copy of Timetable" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = parse_timetable_sheet(ws)
    if not rows:
        raise SystemExit("No rows parsed for 2026-06-01 .. 2026-07-17")

    merge_into_machine(rows)
    print(f"Imported {len(rows)} staff timetable cells from {xlsx.name} ({sheet_name})")

    export_py = OUT / "export_spreadsheet_reference_js.py"
    if export_py.is_file():
        subprocess.run([sys.executable, str(export_py)], check=True, cwd=str(ROOT))
        print("Regenerated working_ui/portal/spreadsheet_reference_data.js")


if __name__ == "__main__":
    main()
